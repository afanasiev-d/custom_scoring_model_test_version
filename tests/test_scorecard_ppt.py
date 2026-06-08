"""Unit tests for scorecard_ppt.py — the formatted Excel export, including the
Confidence-intervals sheet and its 2-decimal number format."""
import io

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

import scorecard_ppt as spt
import viz


# ── helpers ───────────────────────────────────────────────────────────────
def test_safe_sheet_name_sanitises_and_dedupes():
    used = set()
    a = spt._safe_sheet_name("Cash:Flow/2024", used)
    assert ":" not in a and "/" not in a
    b = spt._safe_sheet_name("Cash:Flow/2024", used)       # same input again
    assert a != b                                          # made unique
    long = spt._safe_sheet_name("x" * 80, used)
    assert len(long) <= 31


def test_num_format_two_dp_for_ci_columns():
    s = pd.Series([24.57, 0.67, 56.4])
    assert spt._num_format("Estimate", s) == "#,##0.00"
    assert spt._num_format("CI lower", s) == "#,##0.00"
    assert spt._num_format("Bootstrap SE", s) == "#,##0.00"


def test_num_format_percent_and_whole():
    assert spt._num_format("approval rate", pd.Series([0.1, 0.9])) == "0.00%"
    assert spt._num_format("count", pd.Series([1, 2, 3])) == "#,##0"


# ── full workbook ─────────────────────────────────────────────────────────
@pytest.fixture
def minimal_frames():
    df_scorecard = pd.DataFrame({"Feature": ["Intercept", "f1"], "Category": ["", "A"],
                                 "WoE": [np.nan, 0.5], "Share (%)": [np.nan, 100.0],
                                 "Score": [450.0, 12.0]})
    df_ppt = pd.DataFrame({"cutoff_score": [460, 470], "approval rate": [0.6, 0.4]})
    df_missing = pd.DataFrame({"Feature": ["f1"], "Missing rate": [1.5]})
    df_iv = pd.Series({"f1": 0.21}, name="IV")
    df_ci = pd.DataFrame({
        "Metric": ["KS (0–100)", "AUC ROC (0–1)", "Gini (0–100)"],
        "Estimate": [24.57, 0.67, 34.63], "CI lower": [18.62, 0.63, 25.86],
        "CI upper": [32.94, 0.71, 43.41], "Bootstrap SE": [3.87, 0.02, 4.77],
        "Method": ["95% BCa · stratified · 600 resamples"] * 3,
    })
    return df_scorecard, df_ppt, df_missing, df_iv, df_ci


def test_create_produces_expected_sheets(minimal_frames):
    viz.reset_gallery()
    sc, ppt, miss, iv, ci = minimal_frames
    data = spt.create(sc, ppt, miss, iv, {}, None, df_ci=ci)
    wb = load_workbook(io.BytesIO(data))
    for sheet in ("Scorecard", "Confidence intervals", "PPT", "Missing rate", "Initial IV"):
        assert sheet in wb.sheetnames
    # CI sheet appears right after Scorecard.
    assert wb.sheetnames.index("Confidence intervals") == wb.sheetnames.index("Scorecard") + 1


def test_ci_sheet_values_and_two_dp_format(minimal_frames):
    viz.reset_gallery()
    sc, ppt, miss, iv, ci = minimal_frames
    data = spt.create(sc, ppt, miss, iv, {}, None, df_ci=ci)
    ws = load_workbook(io.BytesIO(data))["Confidence intervals"]
    header = [c.value for c in ws[1]]
    assert header[:5] == ["Metric", "Estimate", "CI lower", "CI upper", "Bootstrap SE"]
    # numeric cells carry the 2-dp format
    for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=5):
        for cell in row:
            assert cell.number_format == "#,##0.00"
    # values round-trip
    assert ws.cell(row=2, column=2).value == pytest.approx(24.57)


def test_create_works_without_ci(minimal_frames):
    viz.reset_gallery()
    sc, ppt, miss, iv, _ = minimal_frames
    data = spt.create(sc, ppt, miss, iv, {}, None)        # df_ci omitted
    wb = load_workbook(io.BytesIO(data))
    assert "Confidence intervals" not in wb.sheetnames
    assert "Scorecard" in wb.sheetnames
